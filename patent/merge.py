from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import glob
import os


# 这里的path是指文件存放的路径，可以把所有的文件放在同一个文件夹，不需要分类
path = '.'

A_class = ['北京大学', '中国人民大学', '清华大学', '北京航空航天大学', '北京理工大学', '中国农业大学', '北京师范大学', '中央民族大学',
           '南开大学', '天津大学', '大连理工大学', '吉林大学', '哈尔滨工业大学', '复旦大学', '同济大学', '上海交通大学', '华东师范大学', '南京大学',
           '东南大学', '浙江大学', '中国科学技术大学', '厦门大学', '山东大学', '中国海洋大学', '武汉大学', '华中科技大学', '中南大学', '中山大学',
           '华南理工大学', '四川大学', '重庆大学', '电子科技大学', '西安交通大学', '西北工业大学',
           '兰州大学', '国防科技大学']

B_class = ['东北大学', '郑州大学', '湖南大学', '云南大学', '西北农林科技大学', '新疆大学']

header = ['序号', '名称', '申请号', '申请日', '申请人', '权利要求数量', '同族国家', '同族数量', '被引证次数', 'IPC分类号', '法律信息']

for k, v in {'A': A_class, 'B': B_class, 'JNU': ['暨南大学']}.items():
    # 这里的SAVE_PATH是最终Excel文件的保存位置
    SAVE_PATH = 'FinalResult-' + k + '.xlsx'
    final_wb = Workbook()
    for university in v:
        s = [header, ]
        final_ws = final_wb.create_sheet(university)
        files = glob.glob(os.path.join(path, k+'-'+university+'*'))
        for f in files:
            wb = load_workbook(f)
            ws = wb['Sheet1']
            for row in list(ws.rows)[1:]:
                r = []
                for cell in row[1:]:
                    r.append(cell.value)
                s.append(r)
        df = pd.DataFrame(s)
        for r in list(dataframe_to_rows(df, header=False))[1:]:
            final_ws.append(r)
    final_wb.save(SAVE_PATH)
