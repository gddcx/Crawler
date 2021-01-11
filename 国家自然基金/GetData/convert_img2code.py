import glob
import os
from aip import AipOcr
import cv2 as cv
import re
from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()

APP_ID = '19667692'
APP_KEY = 'AIbvnvmUqPiNTffXKUm5sHwD'
Secret_KEY = 'AonbakRPyhpz8xA8RBbNOdRE2rtwUSoT'

client = AipOcr(APP_ID, APP_KEY, Secret_KEY)

# for year in ['2015', '2016', '2017', '2018', '2019', '2020']:
#     rawpic_path = os.path.join('D:\Project', year, 'raw', '*')
#     save_path = os.path.join('D:\Project', year, 'process')
#     if not os.path.exists(save_path):
#         os.makedirs(save_path)
#     paths = glob.glob(rawpic_path)
#     for path in paths:
#         filename = os.path.basename(path)
#         img = cv.imread(path, cv.IMREAD_GRAYSCALE)
#         x, y = img.shape
#         left = img[:, :y//2]
#         right = img[:, y//2:]
#         cv.imwrite(os.path.join(save_path, filename.split('.')[0] + "_01.jpg"), left)
#         cv.imwrite(os.path.join(save_path, filename.split('.')[0] + "_02.jpg"), right)

pattern_num = re.compile(r'\d+')
pattern_word = re.compile(r'\D+')

def use_ocr_basicGeneral(client, reader):
    out = client.basicGeneral(reader)
    if 'error_code' in out.keys():
        out = use_ocr_basicGeneral(client, reader)
    return out

def use_ocr_basicAccurate(client, reader):
    out = client.basicAccurate(reader)
    if 'error_code' in out.keys():
        out = use_ocr_basicAccurate(client, reader)
    return out

for year in ['2017', '2018', '2019', '2020']:
    ws = wb.create_sheet(year)
    result_list = []
    processimg_path = os.path.join(r'D:\Project', year, 'process', '*')
    # 某一年的所有图片
    paths = glob.glob(processimg_path)
    for path in paths:
        with open(path, 'rb') as f:
            reader = f.read()
            out = use_ocr_basicAccurate(client, reader)
            # out = use_ocr_basicGeneral(client, reader)
            print(out)
            # 'words_result': [{'words': 'A01数学'}, {'words': 'A0101数论'}, {'words': 'A010101解析数论'}
            out = out['words_result']
            result_list.extend(out)
    #         result = {'words': 'A01数学'}

    # 处理写入excel
    line_num = 0
    previous_class = ""
    for i, result in enumerate(result_list):
        num_subject = result['words']
        if not pattern_num.search(num_subject):
            continue
        num_subject=num_subject.replace('o', '0')
        num_subject=num_subject.replace('O', '0')
        num_subject=num_subject.replace('i', '1')
        num_subject=num_subject.replace('I', '1')
        num_subject=num_subject.replace('l', '1')
        next_item = " "
        next_next_item = " "
        next_next_next_item = ""
        if i< len(result_list)-1:
            next_item = result_list[i+1]['words']
            next_item=next_item.replace('o', '0')
            next_item=next_item.replace('O', '0')
            next_item=next_item.replace('i', '1')
            next_item=next_item.replace('I', '1')
            next_item=next_item.replace('l', '1')
        if i < len(result_list)-2:
           next_next_item = result_list[i+2]['words']
           next_next_item=next_next_item.replace('o', '0')
           next_next_item=next_next_item.replace('O', '0')
           next_next_item=next_next_item.replace('i', '1')
           next_next_item=next_next_item.replace('I', '1')
           next_next_item=next_next_item.replace('l', '1')
        if i < len(result_list)-3:
            next_next_next_item = result_list[i + 3]['words']
            next_next_next_item=next_next_next_item.replace('o', '0')
            next_next_next_item=next_next_next_item.replace('O', '0')
            next_next_next_item=next_next_next_item.replace('i', '1')
            next_next_next_item=next_next_next_item.replace('I', '1')
            next_next_next_item=next_next_next_item.replace('l', '1')
        if not pattern_num.search(next_item):
            num_subject = num_subject + next_item
            if not pattern_num.search(next_next_item):
                num_subject = num_subject + next_next_item
                if not pattern_num.search(next_next_next_item):
                    num_subject = num_subject + next_next_next_item
        line_num += 1
        num = pattern_num.search(num_subject).group()
        isnot_num = pattern_word.findall(num_subject)
        class_ = ""
        subject = ""
        if isnot_num:
            if len(isnot_num[0]) < 2:
                # A B C D
                class_ = isnot_num[0]
                # 具体名称
                subject = ''.join(isnot_num[1:])
            else:
                class_ = previous_class
                subject = ''.join(isnot_num[0])
        else:
            print(isnot_num, num_subject)
        print([class_+num, subject])
        ws.append([class_+num, subject])
        previous_class = class_
        if len(class_+num)<6:
            ws['A' + str(line_num)].font = Font(bold=True)
            ws['B' + str(line_num)].font = Font(bold=True)

wb.save('code_15_20.xlsx')