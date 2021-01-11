import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
import re
wb = Workbook()


url_dict = {
    '2008':'http://www.nsfc.gov.cn/nsfc/cen/xmzn/2008xmzn/10gj/index.htm',
    '2009': 'http://www.nsfc.gov.cn/nsfc/cen/xmzn/2009xmzn/14/index.htm',
    '2010': 'http://www.nsfc.gov.cn/nsfc/cen/xmzn/2010xmzn/14/index.html',
    '2011': 'http://www.nsfc.gov.cn/nsfc/cen/xmzn/2011xmzn/13/index.html',
    '2012': 'http://www.nsfc.gov.cn/nsfc/cen/xmzn/2012xmzn/14/index.html',
    '2013': 'http://www.nsfc.gov.cn/nsfc/cen/xmzn/2013xmzn/15/index.html',
    '2014': 'http://www.nsfc.gov.cn/nsfc/cen/xmzn/2014xmzn/17/index.html'
}
def get2008():
    ws = wb.create_sheet('2008')
    url = url_dict['2008']
    strhtml = requests.get(url)
    soup = BeautifulSoup(strhtml.content, 'html5lib')
    # lxml解析器匹配不了长标签
    data = soup.select('a.kjkx')

    link_list = []
    for d in data:
        l = d.get('href')
        link_list.append('/'.join(url.split('/')[:-1]) + '/' + l)
        # link_list.append(url.replace('index.htm?', l))
    link_list2 = list(set(link_list))
    link_list2.sort(key=link_list.index)

    count = 1
    previous_title = None
    for ul in link_list2:
        strhtml = requests.get(ul)
        # strhtml.encoding = strhtml.apparent_encoding
        # print(strhtml.text)
        soup = BeautifulSoup(strhtml.content, "html5lib")
        title = soup.select('table> tbody > tr > td > table > tbody > tr> td > div > table > tbody > tr > td > p> b')
        if title!=[]:
            current_title = title[0].get_text().strip()
            if previous_title != current_title:
                # 先填入文字，再改变字体格式
                ws.append([current_title])
                ws["A" + str(count)].font = Font(bold=True)
                count += 1
                previous_title = current_title
        data = soup.select('table > tbody > tr > td > div > table > tbody > tr > td')
        # print(type(data))
        temp_result = []
        for d in data[1:]:
            # print(d.get_text().strip())
            # print('-'*10)
            item = d.get_text().strip()
            if item == '':
                continue
            temp_result.append(item)
            if len(temp_result) == 2:
                # print(temp_result[1], temp_result[0])
                # if year!='2008':
                #     if len(temp_result[0])<6:
                #         ws.append([temp_result[0], temp_result[1]])
                #         ws['A' + str(count)].font = Font(bold=True)
                #         ws['B' + str(count)].font = Font(bold=True)
                #     else:
                #         ws.append([temp_result[0], temp_result[1]])
                # else:
                if len(temp_result[1])<6:
                    ws.append([temp_result[1], temp_result[0]])
                    ws['A' + str(count)].font = Font(bold=True)
                    ws['B' + str(count)].font = Font(bold=True)
                else:
                    ws.append([temp_result[1], temp_result[0]])
                temp_result = []
                count += 1

def get2009():
    ws = wb.create_sheet('2009')
    url = url_dict['2009']
    strhtml = requests.get(url)
    soup = BeautifulSoup(strhtml.content, 'html5lib')
    # lxml解析器匹配不了长标签
    data = soup.select('a.kjkx')

    link_list = []
    for d in data:
        l = d.get('href')
        link_list.append('/'.join(url.split('/')[:-1]) + '/' + l)
    link_list2 = list(set(link_list))
    link_list2.sort(key=link_list.index)

    count = 1
    previous_title = None
    for ul in link_list2:
        strhtml = requests.get(ul)
        # strhtml.encoding = strhtml.apparent_encoding
        # print(strhtml.text)
        soup = BeautifulSoup(strhtml.content, "html5lib")
        title = soup.select('table> tbody > tr > td > table > tbody > tr> td > div > table > tbody > tr > td > p> b')
        if title != []:
            current_title = title[0].get_text().strip()
            if previous_title != current_title:
                # 先填入文字，再改变字体格式
                ws.append([current_title])
                ws["A" + str(count)].font = Font(bold=True)
                count += 1
                previous_title = current_title
        data = soup.select('table > tbody > tr > td')
        temp_result = []
        for d in data[1:]:
            item0 = d.get_text().strip()
            item0 = ''.join(re.split(r'[\n\t\s]', item0))
            # print(item0)
            # print('-' * 50)
            for item in item0.split():
                if len(item)>50:
                    continue
                if item == '':
                    continue
                temp_result.append(item)
                if len(temp_result) == 2:
                    if len(temp_result[0])<6:
                        ws.append([temp_result[0], temp_result[1]])
                        ws['A' + str(count)].font = Font(bold=True)
                        ws['B' + str(count)].font = Font(bold=True)
                    else:
                        ws.append([temp_result[0], temp_result[1]])
                    temp_result = []
                    count += 1

def get2010():
    ws = wb.create_sheet('2010')
    url = url_dict['2010']
    strhtml = requests.get(url)
    soup = BeautifulSoup(strhtml.content, 'html5lib')
    # lxml解析器匹配不了长标签
    data = soup.select('a.kjkx')

    link_list = []
    for d in data:
        l = d.get('href')
        link_list.append('/'.join(url.split('/')[:-1]) + '/' + l)
        # link_list.append(url.replace('index.htm?', l))
    link_list2 = list(set(link_list))
    link_list2.sort(key=link_list.index)

    count = 1
    previous_title = None
    for ul in link_list2:
        strhtml = requests.get(ul)
        # strhtml.encoding = strhtml.apparent_encoding
        # print(strhtml.text)
        # print(strhtml.text)
        soup = BeautifulSoup(strhtml.content, "html5lib")
        title = soup.select('body > table > tbody > tr > td> table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr > td > p')
        if title[1] != []:
            current_title = title[1].get_text().strip()
            if previous_title != current_title:
                # 先填入文字，再改变字体格式
                ws.append([current_title])
                ws["A" + str(count)].font = Font(bold=True)
                count += 1
                previous_title = current_title
        data = soup.select('table > tbody > tr > td > p')
        temp_result = []
        for d in data[2:]:
            item0 = d.get_text().strip()
            item0 = ''.join(re.split(r'[\n\t]', item0))
            if len(item0) > 30:
                continue
            for item in item0.split():
                if item == ['']:
                    continue
                temp_result.append(item)
                if len(temp_result) == 2:
                    # print(temp_result[0], temp_result[1])
                    if len(temp_result[0])<6:
                        ws.append([temp_result[0], temp_result[1]])
                        ws['A' + str(count)].font = Font(bold=True)
                        ws['B' + str(count)].font = Font(bold=True)
                    else:
                        ws.append([temp_result[0], temp_result[1]])

                    temp_result = []
                    count += 1

def get2011():
    ws = wb.create_sheet('2011')
    url = url_dict['2011']
    strhtml = requests.get(url)
    soup = BeautifulSoup(strhtml.content, 'html5lib')
    # lxml解析器匹配不了长标签
    data = soup.select('a.jjyw')

    link_list = []
    for d in data:
        l = d.get('href')
        link_list.append('/'.join(url.split('/')[:-1]) + '/' + l)
        # link_list.append(url.replace('index.htm?', l))
    link_list2 = list(set(link_list))
    link_list2.sort(key=link_list.index)

    count = 1
    previous_title = None
    for index, ul in enumerate(link_list2[:2]+link_list[3:7]):
        strhtml = requests.get(ul)
        soup = BeautifulSoup(strhtml.content, "html5lib")
        title = soup.select(
            'body > table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr> td > p.STYLE2')
        if title[0] != []:
            current_title = title[0].get_text().strip()
            if previous_title != current_title:
                # 先填入文字，再改变字体格式
                ws.append([current_title])
                ws["A" + str(count)].font = Font(bold=True)
                count += 1
                previous_title = current_title
        # if index != 2:
        data = soup.select('table > tbody > tr > td > p')
        # else:
        #     data = soup.select('table > tbody > tr > td > * > span')
            # data = soup.select('table>tbody>tr>td>div>table>tbody>tr>td>p.MsoNormal')
        temp_result = []
        item_len = 2
        for d in data[1:]:
            item = d.get_text().strip()
            if item=='':
                continue
            print(item)
            print('-'*50)
            if 'E' in item or 'F' in item:
                if 'E050402' in item or 'E050504' in item or 'E050503' in item or 'E050802' in item or 'E050901' in item or\
                        'E070202' in item or 'E080202' in item or 'E080406' in item or 'E080507' in item :
                    item_len = 3
                elif 'F010205' in item or 'F010802' in item or 'F010803' in item or 'F010804' in item  or 'F010806' in item or \
                        'F010807' in item or 'F010808' in item or 'F010806' in item or 'F010809' in item or 'F010810' in item:
                    item_len = 3
            if item!=[]:
                if len(item)>50:
                    continue
                temp_result.append(item)
                if len(temp_result) == item_len:
                    temp_result[1] = ''.join(temp_result[1:])
                    # print(temp_result[0],temp_result[1])
                    if len(temp_result[0]) < 6:
                        ws.append([temp_result[0],temp_result[1]])
                        ws['A' + str(count)].font = Font(bold=True)
                        ws['B' + str(count)].font = Font(bold=True)
                    else:
                        ws.append([temp_result[0],temp_result[1]])
                    item_len = 2
                    temp_result = []
                    count += 1
    for index, ul in enumerate([link_list2[2], link_list2[-1]]):
        strhtml = requests.get(ul)
        soup = BeautifulSoup(strhtml.content, "html5lib")
        title = soup.select(
            'body > table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr> td > p.STYLE2')
        if title[0] != []:
            current_title = title[0].get_text().strip()
            if previous_title != current_title:
                # 先填入文字，再改变字体格式
                ws.append([current_title])
                ws["A" + str(count)].font = Font(bold=True)
                count += 1
                previous_title = current_title
        if index != 2:
            data = soup.select('table > tbody > tr > td > p')
        else:
            data = soup.select('table>tbody>tr>td>div>table>tbody>tr>td>p.MsoNormal')
        temp_result = []
        item_len = 2
        for d in data[1:]:
            item = d.get_text().strip()
            if item == '':
                continue
            temp_result = item.split()
            if len(item) > 50:
                continue
            if len(temp_result) != 2:
                print(temp_result)
            if len(temp_result) == item_len:
                # print(temp_result[0],temp_result[1])
                if len(temp_result[0]) < 6:
                    ws.append([temp_result[0], temp_result[1]])
                    ws['A' + str(count)].font = Font(bold=True)
                    ws['B' + str(count)].font = Font(bold=True)
                else:
                    ws.append([temp_result[0], temp_result[1]])
                item_len = 2
                temp_result = []
                count += 1

def get2012():
    ws = wb.create_sheet('2012')
    url = url_dict['2012']
    strhtml = requests.get(url)
    soup = BeautifulSoup(strhtml.content, 'html5lib')
    # lxml解析器匹配不了长标签
    data = soup.select('a.jjyw')

    link_list = []
    for d in data:
        l = d.get('href')
        link_list.append('/'.join(url.split('/')[:-1]) + '/' + l)
    link_list2 = list(set(link_list))
    link_list2.sort(key=link_list.index)

    count = 1
    previous_title = None
    for index, ul in enumerate(link_list2):
        strhtml = requests.get(ul)
        # strhtml.encoding = strhtml.apparent_encoding
        # print(strhtml.text)
        # print(strhtml.text)
        soup = BeautifulSoup(strhtml.content, "html5lib")
        title = soup.select(
            'body > table > tbody > tr > td> table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr > td > p')
        if title[0] != []:
            current_title = title[0].get_text().strip()
            if previous_title != current_title:
                # 先填入文字，再改变字体格式
                ws.append([current_title])
                ws["A" + str(count)].font = Font(bold=True)
                count += 1
                previous_title = current_title
        if index !=2:
            data = soup.select('tr> td > table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr > td')
        else:
            data = soup.select(
                'tr> td > table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr > td > p')
        # print(data)
        temp_result = []
        item_len = 2
        for d in data:
            item0 = d.get_text().strip()
            item0 = item0.split()
            print(item0)
            if 'E050402' in item0 or 'E070202' in item0 or 'E080202' in item0:
                item_len = 3
            for item in item0:
                if item == ['']:
                    continue
                temp_result.append(item)
                if len(temp_result) == item_len:
                    temp_result[1] = ''.join(temp_result[1:])
                    # print(temp_result[0], temp_result[1])
                    if len(temp_result[0]) < 6:
                        ws.append([temp_result[0], temp_result[1]])
                        ws['A' + str(count)].font = Font(bold=True)
                        ws['B' + str(count)].font = Font(bold=True)
                    else:
                        ws.append([temp_result[0], temp_result[1]])
                    item_len = 2
                    temp_result = []
                    count += 1

def get2013():
    ws = wb.create_sheet('2013')
    url = url_dict['2013']
    strhtml = requests.get(url)
    soup = BeautifulSoup(strhtml.content, 'html5lib')
    # lxml解析器匹配不了长标签
    data = soup.select('a.jjyw')

    link_list = []
    for d in data:
        l = d.get('href')
        link_list.append('/'.join(url.split('/')[:-1]) + '/' + l)
    link_list2 = list(set(link_list))
    link_list2.sort(key=link_list.index)

    count = 1
    previous_title = None
    for index, ul in enumerate(link_list2):
        strhtml = requests.get(ul)
        # strhtml.encoding = strhtml.apparent_encoding
        # print(strhtml.text)
        # print(strhtml.text)
        soup = BeautifulSoup(strhtml.content, "html5lib")
        title = soup.select(
            'body > table > tbody > tr > td > table > tbody > tr> td > table > tbody > tr > td> table > tbody > tr> td > table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr > td > p')
        if title[0] != []:
            current_title = title[0].get_text().strip()
            if previous_title != current_title:
                # 先填入文字，再改变字体格式
                ws.append([current_title])
                ws["A" + str(count)].font = Font(bold=True)
                count += 1
                previous_title = current_title
        data = soup.select('body > table > tbody > tr > td> table > tbody > tr> td > table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr> td > table > tbody > tr > td > table > tbody > tr > td > p')
        # print(data)
        temp_result = []
        item_len = 2
        for d in data:
            item0 = d.get_text().strip()
            item0 = item0.split()
            for item in item0:
                temp_result.append(item)
                if len(temp_result) == item_len:
                    temp_result[1] = ''.join(temp_result[1:])
                    # print(temp_result[0], temp_result[1])
                    if len(temp_result[0]) < 6:
                        ws.append([temp_result[0], temp_result[1]])
                        ws['A' + str(count)].font = Font(bold=True)
                        ws['B' + str(count)].font = Font(bold=True)
                    else:
                        ws.append([temp_result[0], temp_result[1]])
                    temp_result = []
                    count += 1

def get2014():
    ws = wb.create_sheet('2014')
    url = url_dict['2014']
    strhtml = requests.get(url)
    soup = BeautifulSoup(strhtml.content, 'html5lib')
    # lxml解析器匹配不了长标签
    data = soup.select('a.jjyw')

    link_list = []
    for d in data:
        l = d.get('href')
        link_list.append('/'.join(url.split('/')[:-1]) + '/' + l)
    link_list2 = list(set(link_list))
    link_list2.sort(key=link_list.index)

    count = 1
    previous_title = None
    for index, ul in enumerate(link_list2):
        strhtml = requests.get(ul)
        soup = BeautifulSoup(strhtml.content, "html5lib")
        title = soup.select(
            'body > table > tbody > tr > td> table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr > td > p')
        print(title)
        if title[0] != []:
            current_title = title[0].get_text().strip()
            if previous_title != current_title:
                # 先填入文字，再改变字体格式
                ws.append([current_title])
                ws["A" + str(count)].font = Font(bold=True)
                count += 1
                previous_title = current_title
        data = soup.select('body > table > tbody > tr > td > table > tbody > tr> td > table > tbody > tr > td > table > tbody > tr> td > table > tbody > tr > td > table > tbody > tr > td > table > tbody > tr > td > p')
        # print(data)
        temp_result = []
        item_len = 2
        for d in data:
            item0 = d.get_text().strip()
            item0 = item0.split()
            if len(item0)>2:
                item_len = len(item0)
            for item in item0:
                temp_result.append(item)
                if len(temp_result) == item_len:
                    temp_result[1] = ''.join(temp_result[1:])
                    # print(temp_result[0], temp_result[1])
                    if len(temp_result[0]) < 6:
                        ws.append([temp_result[0], temp_result[1]])
                        ws['A' + str(count)].font = Font(bold=True)
                        ws['B' + str(count)].font = Font(bold=True)
                    else:
                        ws.append([temp_result[0], temp_result[1]])
                    item_len=2
                    temp_result = []
                    count += 1
# ok
get2008()
# ok B部分有个地方要手动 A04...
get2009()
# ok
get2010()
# 手动一下A040509, 手动补 C0508 '生物物理、生物化学与分子生物学研究的新方法与新技术'
get2011()
# ok E0902手动
get2012()
#  ok
get2013()
# ok
get2014()
wb.save('code_08_14.xlsx')
