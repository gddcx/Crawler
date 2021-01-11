# 旧版本老板不满意，老板提了新的要求
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

work_book = load_workbook('../GetData/申请代码.xlsx')

work_book_result = Workbook()

for year in range(2008, 2020):
    ws = work_book_result.create_sheet('%s-%s' % (year, year+1))
    ws['E1'] = '代码、内容有无变化（Yes/No)'
    ws['F1'] = '代码变化(Yes/No)'
    ws['G1'] = '原有代码、内容变化 （原内容）'
    ws['H1'] = '内容变化（Yes/No)'
    ws['I1'] = '原有内容删除/移动'
    ws['J1'] = '当前内容来源'
    index_original_sheet_pre = 0
    index_original_sheet_next = 0
    index_target_sheet_row = 2

    sheet_previous = work_book[str(year)]
    sheet_next = work_book[str(year+1)]

    column_pre_code = [code.value for code in sheet_previous['A']]+[' ']
    column_pre_name = [name.value for name in sheet_previous['B']]+[' ']
    column_next_code = [code.value for code in sheet_next['A']]+[' ']
    column_next_name = [name.value for name in sheet_next['B']]+[' ']

    while(1):
        if index_original_sheet_pre==len(column_pre_name)-1 and index_original_sheet_next==len(column_next_name)-1:
            break
        try:
            # 代码内容均无变化
            if column_pre_code[index_original_sheet_pre] == column_next_code[index_original_sheet_next] \
                    and column_pre_name[index_original_sheet_pre] == column_next_name[index_original_sheet_next]:
                ws['E'+str(index_target_sheet_row)] = 'No'
                ws['A' + str(index_target_sheet_row)] = column_next_code[index_original_sheet_next]
                ws['B' + str(index_target_sheet_row)] = column_next_name[index_original_sheet_next]
                if len(column_next_code[index_original_sheet_next]) < 6:
                    ws['A' + str(index_target_sheet_row)].font = Font(bold=True)
                    ws['B' + str(index_target_sheet_row)].font = Font(bold=True)
                index_original_sheet_pre+=1
                index_original_sheet_next+=1
            # 代码无变化，内容变化
            elif column_pre_code[index_original_sheet_pre] == column_next_code[index_original_sheet_next] \
                    and column_pre_name[index_original_sheet_pre] != column_next_name[index_original_sheet_next]:
                ws['E' + str(index_target_sheet_row)] = 'Yes'
                ws['F' + str(index_target_sheet_row)] = 'No'
                ws['H' + str(index_target_sheet_row)] = 'Yes'
                # 当前内容原本对应其他代码
                if column_next_name[index_original_sheet_next] in column_pre_name:
                    id = column_pre_name.index(column_next_name[index_original_sheet_next])
                    # 当前name对应上一年的什么代码
                    in_pre_code = column_pre_code[id]
                    in_pre_name = column_pre_name[id]
                    ws['J' + str(index_target_sheet_row)] = in_pre_code + '->'
                    # 当前code原本的name是什么
                    id = column_pre_code.index(column_next_code[index_original_sheet_next])
                    pre_name = column_pre_name[id]
                    # 原内容还在
                    if pre_name in column_next_name:
                        id = column_next_name.index(pre_name)
                        ws['I' + str(index_target_sheet_row)] = pre_name + '->' + column_next_code[id]
                    # 原内容删掉了
                    else:
                        ws['I' + str(index_target_sheet_row)] = '-'+pre_name
                # 当前内容是新增的，旧内容被删除了
                else:
                    # 当前code原本的name是什么
                    id = column_pre_code.index(column_next_code[index_original_sheet_next])
                    pre_name = column_pre_name[id]
                    # 原内容还在
                    if pre_name in column_next_name:
                        id = column_next_name.index(pre_name)
                        ws['I' + str(index_target_sheet_row)] = pre_name + '->' + column_next_code[id]
                    # 原内容删掉了
                    else:
                        ws['I' + str(index_target_sheet_row)] = '-' + pre_name
                ws['A' + str(index_target_sheet_row)] = column_next_code[index_original_sheet_next]
                ws['B' + str(index_target_sheet_row)] = column_next_name[index_original_sheet_next]
                if len(column_next_code[index_original_sheet_next]) < 6:
                    ws['A' + str(index_target_sheet_row)].font = Font(bold=True)
                    ws['B' + str(index_target_sheet_row)].font = Font(bold=True)
                index_original_sheet_pre += 1
                index_original_sheet_next += 1
            # 代码删除
            elif not column_pre_code[index_original_sheet_pre] in column_next_code:
                ws['E' + str(index_target_sheet_row)] = 'Yes'
                ws['F' + str(index_target_sheet_row)] = 'Yes'
                ws['G' + str(index_target_sheet_row)] = column_pre_code[index_original_sheet_pre]
                ws['H' + str(index_target_sheet_row)] = 'Yes'
                id = column_pre_code.index(column_pre_code[index_original_sheet_pre])
                pre_name = column_pre_name[id]
                # 内容不是被删除，而是被移到了别处
                if pre_name in column_next_name:
                    id = column_next_name.index(pre_name)
                    next_code = column_next_code[id]
                    ws['I' + str(index_target_sheet_row)] = pre_name + '->' + next_code
                # 内容被删除了
                else:
                    ws['I' + str(index_target_sheet_row)] = '-' + pre_name
                # ws['A' + str(index_target_sheet_row)] = ''
                # ws['B' + str(index_target_sheet_row)] = ''

                index_original_sheet_pre += 1
            # 代码新增
            elif not column_next_code[index_original_sheet_next] in column_pre_code:
                ws['E' + str(index_target_sheet_row)] = 'Yes'
                ws['F' + str(index_target_sheet_row)] = 'Yes'

                ws['H' + str(index_target_sheet_row)] = 'Yes'
                # 内容不是新增，是改变了代码
                if column_next_name[index_original_sheet_next] in column_pre_name:
                    id = column_pre_name.index(column_next_name[index_original_sheet_next])
                    pre_code = column_pre_code[id]
                    pre_name = column_pre_name[id]
                    ws['J' + str(index_target_sheet_row)] = pre_code + '->'
                ws['A' + str(index_target_sheet_row)] = column_next_code[index_original_sheet_next]
                ws['B' + str(index_target_sheet_row)] = column_next_name[index_original_sheet_next]
                if len(column_next_code[index_original_sheet_next]) < 6:
                    ws['A' + str(index_target_sheet_row)].font = Font(bold=True)
                    ws['B' + str(index_target_sheet_row)].font = Font(bold=True)
                index_original_sheet_next += 1
            index_target_sheet_row +=1
        except IndexError:
            print(year, index_target_sheet_row, index_original_sheet_pre, index_original_sheet_next)
work_book_result.save('result.xlsx')

