from openpyxl import load_workbook, Workbook
import re
import difflib
import itertools

work_book = load_workbook('../GetData/申请代码.xlsx')
work_book_result = Workbook()

def get_code_and_name(year):
    """
    统计各个等级下代码的数量变化
    :param begin_year: 起始年份
    :param end_year: 终结年份
    :param faculty: 学部
    :param code_leval: 代码等级

    return {'the_first_level_code': the_first_level_code, 'the_second_level_code':the_second_level_code,
                'the_third_level_code':the_third_level_code, 'the_first_level_name':the_first_level_name,
                'the_second_level_name': the_second_level_name, 'the_third_level_name':the_third_level_name}
    """

    col_code = work_book[str(year)]['A']
    # 获取一级代码
    the_first_level_code = [c.value for c in col_code if len(c.value)==3]

    # 获取二级代码
    the_second_level_code = [[] for _ in range(len(the_first_level_code))]
    first_code_index = 0
    first_code = the_first_level_code[first_code_index]
    for c in col_code:
        if first_code in c.value and len(c.value)==5:
            the_second_level_code[first_code_index].append(c.value)
        elif not first_code in c.value and len(c.value)==5:
            first_code_index+=1
            first_code = the_first_level_code[first_code_index]
            the_second_level_code[first_code_index].append(c.value)

    # 获取三级代码
    the_third_level_code = [[] for _ in range(len(the_first_level_code))]
    first_code_index = 0
    second_code_index = 0
    third_code_index = 0
    second_level = the_second_level_code[first_code_index]
    second_code = second_level[second_code_index]
    temp_third_level = [[] for _ in range(len(second_level))]
    for i, c in enumerate(col_code[2:]):
        # 处理EF识别错误的数据
        if i+3<len(col_code) and col_code[2+i-1].value[0] == col_code[2+i+1].value[0] \
                and col_code[2+i].value[0] != col_code[2+i-1].value[0]:
            col_code[2 + i].value = col_code[2+i-1].value[0] + col_code[2 + i].value[1:]
        # 三级代码
        if second_code in c.value and len(c.value)==7:
            temp_third_level[second_code_index].append(c.value)
        # 二级代码
        elif not second_code in c.value and len(c.value)==5:
            second_code_index += 1
            if second_code_index==len(second_level):
                first_code_index+=1
                second_level = the_second_level_code[first_code_index]
                second_code_index=0
                the_third_level_code[third_code_index] = temp_third_level
                third_code_index+=1
                temp_third_level = [[] for _ in range(len(second_level))]
            second_code = second_level[second_code_index]
            if len(second_code) == 6 or len(second_code) == 4:
                second_code = second_code + str(int(second_level[second_code_index-1][-1]) + 1)
    the_third_level_code[third_code_index] = temp_third_level

    # print(the_first_level_code)
    # print(the_second_level_code)
    # print(the_third_level_code)

    row_index = 0
    col_name =  work_book[str(year)]['B']
    the_first_level_name = []
    the_second_level_name = []
    the_third_level_name = []
    for first_index in range(len(the_first_level_code)):
        the_first_level_name.append(col_name[row_index].value)
        row_index += 1
        temp_the_second_name = []
        temp_the_third_name1 = []
        for second_index in range(len(the_second_level_code[first_index])):
            temp_the_second_name.append(col_name[row_index].value)
            row_index += 1
            temp_the_third_name2 = []
            if the_third_level_code[first_index][second_index] == []:
                pass
            else:
                for third_index in range(len(the_third_level_code[first_index][second_index])):
                    temp_the_third_name2.append(col_name[row_index].value)
                    row_index+=1
            temp_the_third_name1.append(temp_the_third_name2)
        the_third_level_name.append(temp_the_third_name1)
        the_second_level_name.append(temp_the_second_name)

    # print(the_first_level_name)
    # print(the_second_level_name)
    # print(the_third_level_name)
    return {'the_first_level_code': the_first_level_code, 'the_second_level_code':the_second_level_code,
            'the_third_level_code':the_third_level_code, 'the_first_level_name':the_first_level_name,
            'the_second_level_name': the_second_level_name, 'the_third_level_name':the_third_level_name}

def statistic(data_first_year_ori, data_second_year_ori, first_year_index=None, second_year_index=None, level='first'):
    if first_year_index !=None and second_year_index != None:
        data_first_year = data_first_year_ori['the_' + level + '_level_name']
        data_second_year = data_second_year_ori['the_' + level + '_level_name']
        data_first_year_code = data_first_year_ori['the_' + level + '_level_code']
        data_second_year_code = data_second_year_ori['the_' + level + '_level_code']
        for i, j in zip(first_year_index, second_year_index):
            data_first_year = data_first_year[i]
            data_second_year = data_second_year[j]
            data_first_year_code = data_first_year_code[i]
            data_second_year_code = data_second_year_code[j]
    else:
        data_first_year = data_first_year_ori['the_'+level+'_level_name']
        data_second_year = data_second_year_ori['the_'+level+'_level_name']
        data_first_year_code = data_first_year_ori['the_' + level + '_level_code']
        data_second_year_code = data_second_year_ori['the_' + level + '_level_code']
    if data_first_year == [] and data_second_year ==[]:
        return {}, {}
    add = {}
    delete = {}
    # 获取增删情况
    # 增加的情况
    if data_second_year != []:
        for i, name in enumerate(data_second_year):
            if not name in data_first_year:
                add[name] = data_second_year_code[i]
    # 删除的情况
    if data_first_year != []:
        for i, name in enumerate(data_first_year):
            if not name in data_second_year:
                delete[name] = data_first_year_code[i]
    # 处理同一个项目改了名字的情况
    # TODO:字符串相似度匹配，判断是不是改了名字，2016 医学病原微生物与感染, 2017 医学病原生物与感染，这个应该也要作为输出？
    # the_same_item = []
    # for name_add, i in add.items():
    #     for name_del, j in delete.items():
    #         if difflib.SequenceMatcher(None, name_add, name_del).quick_ratio()>0.9:
    #             data_first_year[j] = name_add
    #             # 从这里获取改名的情况
    #             the_same_item.append([name_add, name_del])
    # for name_add, name_del in the_same_item:
    #     add.pop(name_add)
    #     delete.pop(name_del)
    # print(first_level_add, first_level_delete)
    # print(data_first_year['the_first_level_name'])
    # 两年的索引映射
    mapping = {}
    if level != 'third':
        for i, name in enumerate(data_second_year):
            if not name in add.keys():
                mapping[str(i)] = data_first_year.index(name)
    # 输出结果
    # if not (add == {} and delete == {}):
    #     print('level:', level, 'add: ', add, 'delete: ', delete)

    # 递归处理下一级代码
    # 第二年的索引是key，第一年的索引是value
    if level == 'first':
        temp_name_code = []
        temp_name_code_second = []
        for key, value in mapping.items():
            name_code, a, d = statistic(data_first_year_ori, data_second_year_ori,
                             first_year_index= [value], second_year_index=[int(key)], level='second')
            if name_code!=[]:
                temp_name_code_second.append(name_code)
            if not (a == {} and d == {}):
                temp_name_code.append({'add':a, 'del':d})
        return temp_name_code_second, temp_name_code, [{'add':add, 'del':delete}]
    elif level == 'second':
        temp_name_code_third = []
        for key, value in mapping.items():
            a, d = statistic(data_first_year_ori, data_second_year_ori,
                             first_year_index=[*first_year_index,value] , second_year_index=[*second_year_index, int(key)],
                             level='third')
            if not (a == {} and d == {}):
                temp_name_code_third.append({'add': a, 'del': d})
        return temp_name_code_third, add, delete
    else:
        return add, delete


fac = {'A':'数理科学部', 'B':'化学科学部', 'C':'生命科学部','D':'地球科学部','E':'工程与材料科学部',
           'F':'信息科学部','G':'管理科学部','H':'医学科学部'}
ws = {}
for key, value in fac.items():
    ws[key] = work_book_result.create_sheet(value)
def judge(add, delete, level):
    try:
        for add_, del_ in itertools.zip_longest(add.items(), delete.items(), fillvalue=''):
            if add_ == '':
                ws = work_book_result[fac[del_[1][0]]]
                ws.append(['', '', del_[1], del_[0]])
            elif del_=='':
                ws = work_book_result[fac[add_[1][0]]]
                ws.append([add_[1], add_[0], '', ''])
            else:
                ws = work_book_result[fac[add_[1][0]]]
                ws.append([add_[1], add_[0], del_[1], del_[0]])
    except ValueError:
        print(add, delete)
    work_book_result.save('result_2016_2017.xlsx')

if __name__ == '__main__':
    data_first = get_code_and_name(2016)
    data_second = get_code_and_name(2017)
    third, second, first = statistic(data_first, data_second, level='first')
    # difflib.SequenceMatcher(None, str1, str2).quick_ratio()
    # 写入文件
    # 一级代码
    judge(first[0]['add'], first[0]['del'], 'first')
    # 二级代码
    for item in second:
        judge(item['add'], item['del'], level='second')
    # 三级代码
    for i in third:
        for j in i:
            # print(j)
            judge(j['add'], j['del'], level='third')